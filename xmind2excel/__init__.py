from pprint import pprint
from configparser import ConfigParser

import openpyxl
from openpyxl.styles import PatternFill
from xmindparser import xmind_to_dict

DEFAULT_OWNER = ''
DEFAULT_TITLE_LINE = ('模块', '验证功能', '验证场景', '用例标题', '优先级', '预置条件', '测试步骤', '期望结果', '归属人')


def load_xmind(xmind_file: str, sheet_index=0) -> dict:
    data = xmind_to_dict(xmind_file)
    pprint(data)
    return data[sheet_index]


# pprint(data)
def parse_testcases(xmind_data: dict, owner=DEFAULT_OWNER):
    root_topic = xmind_data['topic']
    testcases = []

    for module in root_topic['topics']:
        module_title = module['title']

        for feature in module.get('topics', []):
            feature_title = feature['title']

            for scenario in feature.get('topics', []):
                scenario_title = scenario['title']

                for testcase in scenario.get('topics', []):
                    testcase_title = testcase['title']
                    priority = ''
                    precondition = ''
                    owner = owner

                    steps = []
                    excepted = []

                    note = testcase.get('note')
                    if note:
                        precondition = note  # todo 解析 [预置条件]

                    for index, step in enumerate(testcase.get('topics', [])):
                        step_sn = index + 1
                        step_title = step['title']
                        if 'topics' in step.keys():
                            step_excepted = ' \n'.join([item['title'] for item in step['topics']])
                            excepted.append(f'{step_sn}. {step_excepted}')

                        steps.append(f'{step_sn}. {step_title}')

                    makers = testcase.get('makers')
                    print('markers', makers)
                    if makers:
                        priority_markers = [item for item in makers if 'priority' in item]
                        if priority_markers:
                            priority = priority_markers[0].replace('priority-', 'p')

                    test_steps = ' \n'.join(steps)
                    test_excepted = ' \n'.join(excepted)
                    testcase_data = (
                        module_title, feature_title, scenario_title, testcase_title, priority, precondition, test_steps,
                        test_excepted, owner)
                    testcases.append(testcase_data)
                    print('测试用例', testcase_data)
    return testcases


def write_excel(testcases, output_excel_file: str = 'case.xlsx', title_line: list = DEFAULT_TITLE_LINE):
    excel = openpyxl.Workbook()
    sheet = excel.active
    title_line = title_line
    sheet.append(title_line)

    color = PatternFill('solid', fgColor="FF00FF")
    for i in range(1, len(title_line)+1):
        sheet.cell(1, i).fill = color

    for testcase in testcases:
        sheet.append(testcase)

    excel.save(output_excel_file)


def xmind2excel(xmind_file, excel_file=None, owner=None):
    xmind_data = load_xmind(xmind_file)
    testcases = parse_testcases(xmind_data, owner=owner)
    write_excel(testcases, excel_file)


