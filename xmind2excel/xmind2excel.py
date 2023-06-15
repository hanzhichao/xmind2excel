import openpyxl
from openpyxl.styles import PatternFill
from xmindparser import xmind_to_dict

DEFAULT_OWNER = ''
DEFAULT_TITLE_LINE = (
    '模块', '验证功能', '验证场景', '用例标题', '优先级', '预置条件', '测试步骤', '期望结果', '归属人')


def _load_xmind(xmind_file: str, sheet_index=0) -> dict:
    """
    加载xmind指定sheet数据
    :param xmind_file: xmind文件路径，支持xmind8及xmind zen文件
    :param sheet_index:
    :return:
    """
    data = xmind_to_dict(xmind_file)
    return data[sheet_index]


def _parse_steps_excepted(testcase: dict):
    steps = []
    excepted = []

    for index, step in enumerate(testcase.get('topics', [])):
        step_sn = index + 1
        step_title = step['title']
        if 'topics' in step.keys():
            step_excepted = ' \n'.join([item['title'] for item in step['topics']])
            excepted.append(f'{step_sn}. {step_excepted}')

        steps.append(f'{step_sn}. {step_title}')
    return steps, excepted


def _parse_priority(testcase: dict):
    priority = ''
    makers = testcase.get('makers')
    if makers:
        priority_markers = [item for item in makers if 'priority' in item]
        if priority_markers:
            priority = priority_markers[0].replace('priority-', 'p')
    return priority


def _parse_precondition(testcase: dict):
    precondition = ''
    note = testcase.get('note')
    if note:
        precondition = note  # todo 解析 [预置条件]
    return precondition


def _parse_testcases(xmind_data: dict, owner=DEFAULT_OWNER) -> list:
    """
    解析Xmind用例数据
    :param xmind_data: 由_load_xmind加载得到的XMind数据
    :param owner: 制定所有用例归属人
    :return: 用例列表
    """
    root_topic = xmind_data['topic']
    testcases = []

    for module in root_topic['topics']:
        module_title = module['title']

        for feature in module.get('topics', []):
            feature_title = feature['title']

            for scenario in feature.get('topics', []):
                scenario_title = scenario['title']

                for testcase in scenario.get('topics', []):
                    testcase_title = testcase['title']
                    priority = _parse_priority(testcase)
                    precondition = _parse_precondition(testcase)
                    steps, excepted = _parse_steps_excepted(testcase)

                    test_steps = ' \n'.join(steps)
                    test_excepted = ' \n'.join(excepted)

                    testcase_data = (
                        module_title, feature_title, scenario_title, testcase_title, priority, precondition, test_steps,
                        test_excepted, owner)
                    testcases.append(testcase_data)
    return testcases


def _write_excel(testcases: list, excel_file: str, title_line: list = DEFAULT_TITLE_LINE) -> None:
    """
    将测试用例写入Excel文件
    :param testcases: 由_parse_testcases函数得到的用例列表
    :param excel_file: 要写入的Excel文件路径
    :param title_line: Excel标题行
    :return: excel文件路径
    """
    excel = openpyxl.Workbook()
    sheet = excel.active
    title_line = title_line
    sheet.append(title_line)

    color = PatternFill('solid', fgColor="FF00FF")
    for i in range(1, len(title_line) + 1):
        sheet.cell(1, i).fill = color

    for testcase in testcases:
        sheet.append(testcase)

    excel.save(excel_file)


def xmind2excel(xmind_file: str, excel_file: str = None, owner: str = None) -> str:
    """
    Xmind转Excel
    :param xmind_file: XMind文件路径
    :param excel_file: 输出的Excel文件路径，为None时使用和Xmind同样的名称和路径
    :param owner: 制定用例归属人
    :return 生成的Excel文件路径
    """
    if excel_file is None:
        excel_file = xmind_file.replace('.xmind', '.xlsx')
    xmind_data = _load_xmind(xmind_file)
    testcases = _parse_testcases(xmind_data, owner=owner)
    _write_excel(testcases, excel_file)
    return excel_file


if __name__ == '__main__':
    import sys

    _excel_file = None
    _owner = None

    if len(sys.argv) < 2:
        raise ValueError('缺少XMind文件路径参数')

    _xmind_file = sys.argv[1]
    if len(sys.argv) > 2:
        _excel_file = sys.argv[2]
    if len(sys.argv) > 3:
        _owner = sys.argv[3]
    xmind2excel(_xmind_file, _excel_file, _owner)
